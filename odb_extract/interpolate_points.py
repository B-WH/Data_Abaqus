"""Interpolate exported Abaqus ODB node-field data at requested coordinates."""

from __future__ import print_function

import argparse
import csv
import json
import os
import sys

import numpy as np

TOOL_NAME = "odb_extract.interpolate_points"
METADATA_SCHEMA_VERSION = 1


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Interpolate extracted Abaqus ODB node data at coordinate points."
    )
    parser.add_argument("--data", required=True, help="Input NPZ from Extract_data_ODB.py.")
    parser.add_argument("--metadata", required=True, help="Input metadata JSON.")
    parser.add_argument("--points", required=True, help="CSV/XLSX containing point_id,x,y,z rows.")
    parser.add_argument("--output", required=True, help="Output NPZ path.")
    parser.add_argument("--metadata-output", required=True, help="Output metadata JSON path.")
    parser.add_argument("--fields", nargs="+", default=None, help="Optional field names.")
    parser.add_argument("--neighbors", type=int, default=4, help="Neighbor count for interpolation.")
    parser.add_argument(
        "--exact-tol",
        type=float,
        default=1.0e-9,
        help="Distance tolerance for treating a query point as an exact node hit.",
    )
    return parser.parse_args(argv)


def load_metadata(path):
    with open(path, "r", encoding="utf-8") as stream:
        return json.load(stream)


def ensure_parent_dir(path):
    parent = os.path.dirname(os.path.abspath(path))
    if parent and not os.path.isdir(parent):
        os.makedirs(parent)


def _point_id_text(value, fallback):
    text = "" if value is None else str(value).strip()
    return text or str(fallback)


def _points_from_rows(fieldnames, rows):
    normalized = {
        str(name).strip().lower(): name for name in fieldnames if name is not None
    }
    missing = [name for name in ("x", "y", "z") if name not in normalized]
    if missing:
        raise ValueError("Point file is missing required column(s): {}".format(", ".join(missing)))

    points = []
    for row_index, row in enumerate(rows, start=1):
        x_value = row.get(normalized["x"])
        y_value = row.get(normalized["y"])
        z_value = row.get(normalized["z"])
        if x_value in (None, "") and y_value in (None, "") and z_value in (None, ""):
            continue
        point_id_key = normalized.get("point_id")
        points.append(
            {
                "point_id": _point_id_text(
                    row.get(point_id_key) if point_id_key is not None else None,
                    row_index,
                ),
                "coordinates": np.asarray(
                    [float(x_value), float(y_value), float(z_value)],
                    dtype=float,
                ),
            }
        )
    return points


def _read_csv_query_points(path):
    with open(path, newline="", encoding="utf-8") as stream:
        reader = csv.DictReader(stream)
        return _points_from_rows(reader.fieldnames or [], reader)


def _read_excel_query_points(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("Reading Excel point files requires openpyxl.")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            fieldnames = next(rows)
        except StopIteration:
            raise ValueError("Point Excel file is empty.")
        points = _points_from_rows(
            fieldnames,
            (dict(zip(fieldnames, row)) for row in rows),
        )
        return points, sheet.title
    finally:
        workbook.close()


def read_query_point_file(path):
    extension = os.path.splitext(path)[1].lower()
    if extension == ".csv":
        return _read_csv_query_points(path), {
            "path": os.path.abspath(path),
            "format": "csv",
            "worksheet": None,
        }
    if extension in (".xlsx", ".xlsm"):
        points, worksheet = _read_excel_query_points(path)
        return points, {
            "path": os.path.abspath(path),
            "format": extension.lstrip("."),
            "worksheet": worksheet,
        }
    raise ValueError("Unsupported point file format: {}".format(extension or "<none>"))


def _node_coordinate_lookup(data, metadata):
    if "node_coordinates" not in data:
        raise ValueError(
            "NPZ does not contain node_coordinates. Re-run Extract_data_ODB.py with the updated extractor."
        )
    coordinates = np.asarray(data["node_coordinates"], dtype=float)
    nodes = metadata.get("nodes") or []
    if len(nodes) != len(coordinates):
        raise ValueError(
            "Metadata nodes count ({}) does not match node_coordinates rows ({}).".format(
                len(nodes),
                len(coordinates),
            )
        )
    lookup = {}
    for index, node in enumerate(nodes):
        key = (node.get("instance", ""), int(node.get("label")))
        lookup[key] = coordinates[index]
    return lookup


def _available_fields(metadata):
    field_outputs = metadata.get("field_outputs") or {}
    if field_outputs:
        return sorted(field_outputs.keys())
    return sorted(metadata.get("fields") or [])


def _available_node_fields(metadata):
    field_outputs = metadata.get("field_outputs") or {}
    if not field_outputs:
        return _available_fields(metadata)
    return sorted(
        field_name
        for field_name, field_metadata in field_outputs.items()
        if field_metadata.get("location") == "NODE"
    )


def available_node_sets(metadata):
    return sorted((metadata.get("node_sets") or {}).keys())


def _validate_field(metadata, field_name):
    field_outputs = metadata.get("field_outputs") or {}
    field_metadata = field_outputs.get(field_name)
    if field_metadata is None:
        raise ValueError("Field {} is not present in metadata.".format(field_name))
    if field_metadata.get("location") != "NODE":
        raise ValueError("Field {} is not a node field.".format(field_name))
    return field_metadata


def _coordinates_for_field(field_metadata, coordinate_lookup, allowed_node_keys=None):
    coordinates = []
    labels = []
    indexes = []
    for index, point in enumerate(field_metadata.get("points") or []):
        key = (point.get("instance", ""), int(point.get("node_label")))
        if allowed_node_keys is not None and key not in allowed_node_keys:
            continue
        if key not in coordinate_lookup:
            raise ValueError(
                "Node coordinates are missing for instance {} node {}.".format(
                    key[0],
                    key[1],
                )
            )
        labels.append(int(key[1]))
        coordinates.append(coordinate_lookup[key])
        indexes.append(index)
    if not coordinates:
        raise ValueError("Field has no node points to interpolate.")
    return (
        np.asarray(coordinates, dtype=float),
        np.asarray(labels, dtype=np.int64),
        np.asarray(indexes, dtype=np.int64),
    )


def _selected_node_keys(data, metadata, node_sets):
    if not node_sets:
        return None
    nodes = metadata.get("nodes") or []
    definitions = metadata.get("node_sets") or {}
    selected = set()
    for name in node_sets:
        definition = definitions.get(name)
        if not isinstance(definition, dict):
            raise ValueError("Node set {} is not present in cached metadata.".format(name))
        key = definition.get("indices_key")
        if not key or key not in data:
            raise ValueError("Node set {} membership array is missing.".format(name))
        indices = np.asarray(data[key])
        if indices.ndim != 1 or not np.issubdtype(indices.dtype, np.integer):
            raise ValueError("Node set {} membership array is invalid.".format(name))
        if len(indices) != int(definition.get("member_count", -1)):
            raise ValueError("Node set {} member count does not match its array.".format(name))
        if not len(indices) or int(np.min(indices)) < 0 or int(np.max(indices)) >= len(nodes):
            raise ValueError("Node set {} contains an invalid cached node index.".format(name))
        for index in indices.tolist():
            node = nodes[int(index)]
            selected.add((node.get("instance", ""), int(node.get("label"))))
    return selected


def _neighbor_weights(node_coordinates, query_coordinates, neighbors, exact_tol):
    if neighbors < 1:
        raise ValueError("--neighbors must be at least 1.")
    distances = np.linalg.norm(node_coordinates - query_coordinates, axis=1)
    nearest_index = int(np.argmin(distances))
    if float(distances[nearest_index]) <= exact_tol:
        return (
            np.asarray([nearest_index], dtype=np.int64),
            np.asarray([1.0], dtype=float),
            np.asarray([float(distances[nearest_index])], dtype=float),
            "exact",
        )

    count = min(int(neighbors), len(distances))
    if count == len(distances):
        indices = np.argsort(distances).astype(np.int64)
    else:
        partial = np.argpartition(distances, count - 1)[:count]
        indices = partial[np.argsort(distances[partial])].astype(np.int64)
    selected_distances = distances[indices].astype(float)
    inverse_distances = 1.0 / selected_distances
    weights = inverse_distances / np.sum(inverse_distances)
    return indices, weights, selected_distances, "weighted"


def _weighted_values(values, indices, weights):
    return np.tensordot(values[:, indices, :], weights, axes=([1], [0]))


def _save_npz(output_path, arrays):
    ensure_parent_dir(output_path)
    np.savez_compressed(output_path, **arrays)


def _save_metadata(output_path, metadata):
    ensure_parent_dir(output_path)
    with open(output_path, "w", encoding="utf-8") as stream:
        json.dump(metadata, stream, ensure_ascii=False, indent=2, sort_keys=True)


def _build_point_arrays(
    data,
    metadata,
    query_points,
    requested_fields,
    neighbors,
    exact_tol,
    allowed_node_keys=None,
):
    field_metadata_by_name = {
        field_name: _validate_field(metadata, field_name) for field_name in requested_fields
    }
    coordinate_lookup = _node_coordinate_lookup(data, metadata)
    frequencies = np.asarray(data["frequencies"], dtype=float)
    point_records = [
        {
            "point_id": point["point_id"],
            "coordinates": [float(value) for value in point["coordinates"]],
            "fields": {},
        }
        for point in query_points
    ]
    arrays = {
        "frequencies": frequencies,
        "point_ids": np.asarray([point["point_id"] for point in query_points]),
        "point_coordinates": np.asarray(
            [point["coordinates"] for point in query_points],
            dtype=float,
        ),
    }
    field_outputs = {}

    for field_name in requested_fields:
        field_metadata = field_metadata_by_name[field_name]
        node_coordinates, node_labels, field_indexes = _coordinates_for_field(
            field_metadata,
            coordinate_lookup,
            allowed_node_keys=allowed_node_keys,
        )
        components = field_metadata.get("components") or [
            "component_{}".format(index + 1) for index in range(data["{}_real".format(field_name)].shape[2])
        ]
        real_data = np.asarray(data["{}_real".format(field_name)], dtype=float)[
            :, field_indexes, :
        ]
        imag_data = np.asarray(data["{}_imag".format(field_name)], dtype=float)[
            :, field_indexes, :
        ]
        real_output = np.empty(
            (len(frequencies), len(query_points), len(components)),
            dtype=float,
        )
        imag_output = np.empty(real_output.shape, dtype=float)

        for point_index, point in enumerate(query_points):
            indices, weights, distances, method = _neighbor_weights(
                node_coordinates,
                point["coordinates"],
                neighbors,
                exact_tol,
            )
            real_output[:, point_index, :] = _weighted_values(real_data, indices, weights)
            imag_output[:, point_index, :] = _weighted_values(imag_data, indices, weights)
            field_record = {
                "method": method,
                "neighbor_labels": [int(label) for label in node_labels[indices].tolist()],
                "neighbor_weights": [float(value) for value in weights.tolist()],
                "neighbor_distances": [float(value) for value in distances.tolist()],
            }
            point_records[point_index]["fields"][field_name] = field_record
            if "method" not in point_records[point_index]:
                point_records[point_index].update(field_record)

        real_key = "{}_real".format(field_name)
        imag_key = "{}_imag".format(field_name)
        arrays[real_key] = real_output
        arrays[imag_key] = imag_output
        field_outputs[field_name] = {
            "location": "POINT",
            "component_count": len(components),
            "components": list(components),
            "array_layout": ["frame", "point", "component"],
        }

    return arrays, point_records, field_outputs


def interpolate_files(
    data_path,
    metadata_path,
    points_path,
    output_path,
    metadata_output_path,
    fields=None,
    node_sets=None,
    neighbors=4,
    exact_tol=1.0e-9,
):
    source_metadata = load_metadata(metadata_path)
    requested_fields = list(fields if fields is not None else _available_node_fields(source_metadata))
    query_points, point_input = read_query_point_file(points_path)
    if not query_points:
        raise ValueError("Point file does not contain any coordinate rows.")
    with np.load(data_path) as data:
        allowed_node_keys = _selected_node_keys(data, source_metadata, node_sets)
        arrays, point_records, field_outputs = _build_point_arrays(
            data,
            source_metadata,
            query_points,
            requested_fields,
            neighbors,
            exact_tol,
            allowed_node_keys=allowed_node_keys,
        )
        array_shapes = {name: list(array.shape) for name, array in arrays.items()}

    array_layouts = {
        "frequencies": ["frame"],
        "point_ids": ["point"],
        "point_coordinates": ["point", "coordinate"],
    }
    for field_name in requested_fields:
        real_key = "{}_real".format(field_name)
        imag_key = "{}_imag".format(field_name)
        array_layouts[real_key] = ["frame", "point", "component"]
        array_layouts[imag_key] = ["frame", "point", "component"]
    metadata = {
        "tool": {
            "name": TOOL_NAME,
            "metadata_schema_version": METADATA_SCHEMA_VERSION,
        },
        "source_data": os.path.abspath(data_path),
        "source_metadata": os.path.abspath(metadata_path),
        "point_input": point_input,
        "fields": requested_fields,
        "point_count": len(query_points),
        "points": point_records,
        "array_shapes": array_shapes,
        "array_layouts": array_layouts,
        "field_outputs": field_outputs,
        "interpolation": {
            "neighbors": int(neighbors),
            "exact_tol": float(exact_tol),
            "node_sets": list(node_sets or []),
        },
        "warnings": [],
    }
    _save_npz(output_path, arrays)
    _save_metadata(metadata_output_path, metadata)
    return metadata


def main(argv=None):
    args = parse_args(argv)
    try:
        interpolate_files(
            data_path=args.data,
            metadata_path=args.metadata,
            points_path=args.points,
            output_path=args.output,
            metadata_output_path=args.metadata_output,
            fields=args.fields,
            neighbors=args.neighbors,
            exact_tol=args.exact_tol,
        )
    except (OSError, KeyError, ValueError) as exc:
        print("ERROR: {}".format(exc), file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
